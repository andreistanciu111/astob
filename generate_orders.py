# generate_orders.py
# ============================================================
# Citire ASTOB + TABEL CHEIE, completare șablon Excel și
# generare ZIP cu ordinele de plată.
# Păstrează layout-ul din șablon și formatează sumele cu 2 zecimale.
# ============================================================

from __future__ import annotations
import argparse
import io
import os
import zipfile
from datetime import datetime
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# ---------------- Utils de citire ----------------

def _sniff_read(path: str) -> pd.DataFrame:
    """
    Citește atât .xlsx cât și .csv (cu separatori & encodări frecvente).
    """
    name = os.path.basename(path).lower()
    # Excel direct
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xls"):
        return pd.read_excel(path, engine="openpyxl")
    # CSV – încercări tolerate
    for enc in ("utf-8-sig", "cp1250", "cp1252", "latin-1"):
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding=enc)
        except Exception:
            continue
    # Ultima șansă: delimitor comun
    return pd.read_csv(path, sep=";", engine="python", encoding_errors="ignore")


def read_table(path: str) -> pd.DataFrame:
    df = _sniff_read(path)
    # normalizează header-ele: fără spații duble, strip
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ---------------- Normalizări coloane ----------------

def _norm(s: str) -> str:
    return (
        s.lower()
        .replace("ă", "a").replace("â", "a").replace("ș", "s").replace("ţ", "t").replace("ț", "t").replace("î", "i")
        .replace(".", " ").replace("_", " ").replace("-", " ").replace("  ", " ").strip()
    )


def find_col(df: pd.DataFrame, candidates: list[str]) -> str:
    """
    Găsește prima coloană existentă în DataFrame, tolerantă la diacritice/spații/punctuație.
    """
    norms = {_norm(c): c for c in df.columns}
    for cand in candidates:
        nc = _norm(cand)
        if nc in norms:
            return norms[nc]
    # fallback: conține
    for cand in candidates:
        nc = _norm(cand)
        for k, v in norms.items():
            if nc in k:
                return v
    raise KeyError(f"Missing columns: tried {candidates} in {list(df.columns)}")


# ---------------- Conversii & formate ----------------

def parse_amount(v) -> float:
    """
    Transformă valori românești '1.234,56' / '46,16' / '46' în float.
    """
    if pd.isna(v):
        return 0.0
    s = str(v).strip()
    # dacă e deja numeric
    try:
        return float(s)
    except Exception:
        pass
    # curățare românească
    s = s.replace(" ", "")
    if "," in s and "." in s:
        # probabil mii cu punct și zecimale cu virgulă
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    # altfel rămâne așa
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_datetime(date_str, time_str: Optional[str]) -> datetime:
    """
    Din 'Data tranzacției' + 'Ora tranzacției' returnează datetime (inclusiv ora).
    Dacă data conține deja ora, ignoră time_str.
    """
    if pd.isna(date_str):
        return datetime.min

    # Dacă data e deja datetime
    if isinstance(date_str, datetime):
        base = date_str
    else:
        s = str(date_str).strip()
        # încearcă formatele comune
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                base = datetime.strptime(s, fmt)
                break
            except Exception:
                base = None
        if base is None:
            # încearcă pandas
            try:
                base = pd.to_datetime(s, dayfirst=True, errors="coerce")
                if pd.isna(base):
                    return datetime.min
                base = pd.Timestamp(base).to_pydatetime()
            except Exception:
                return datetime.min

    # Dacă data NU are ora și avem coloană separată pentru oră
    if time_str and not (base.hour or base.minute or base.second):
        ts = str(time_str).strip()
        for tfmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = datetime.strptime(ts, tfmt)
                base = base.replace(hour=t.hour, minute=t.minute, second=t.second)
                break
            except Exception:
                pass

    return base


def ro_month_upper(dt: datetime) -> str:
    months = {
        1: "IANUARIE", 2: "FEBRUARIE", 3: "MARTIE", 4: "APRILIE",
        5: "MAI", 6: "IUNIE", 7: "IULIE", 8: "AUGUST",
        9: "SEPTEMBRIE", 10: "OCTOMBRIE", 11: "NOIEMBRIE", 12: "DECEMBRIE",
    }
    return f"{dt.day} {months.get(dt.month, '').upper()} {dt.year}"


# ---------------- Helpers pentru șablon ----------------

def find_cell(ws: Worksheet, needle: str) -> Optional[Tuple[int, int]]:
    """Caută exact textul într-o celulă și întoarce (row, col) 1-based."""
    for r in ws.iter_rows(values_only=False):
        for c in r:
            if isinstance(c.value, str) and c.value.strip() == needle:
                return c.row, c.column
    return None


def replace_placeholder(ws: Worksheet, placeholder: str, value: str):
    """Înlocuiește placeholder-ul oriunde apare într-o celulă de tip string."""
    for r in ws.iter_rows(values_only=False):
        for c in r:
            if isinstance(c.value, str) and placeholder in c.value:
                c.value = c.value.replace(placeholder, value)


def set_col_width(ws: Worksheet, col_idx: int, width: float):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------- Generator per client ----------------

def fill_workbook_for_client(
    template_path: str,
    client_name: str,
    client_rc: str,
    client_cui: str,
    client_addr: str,
    tx_rows: list[dict],
    colectari_text: str,
    header_date_text: str,
    out_path: str,
):
    """
    Copiază șablonul și scrie datele pentru un client.
    tx_rows: listă de dict-uri cu chei: site, tid, produs, valoare (float), data (datetime)
    """
    wb = load_workbook(template_path)
    ws = wb.active  # presupunem un singur sheet: „Foaie1”

    # 1) antet
    replace_placeholder(ws, "{NUME}", str(client_name) if client_name else "")
    replace_placeholder(ws, "{NR. INREGISTRARE R.C.}", str(client_rc) if client_rc else "")
    replace_placeholder(ws, "{CUI}", str(client_cui) if client_cui else "")
    replace_placeholder(ws, "{ADRESA}", str(client_addr) if client_addr else "")
    replace_placeholder(ws, "{HEADER_DATE}", header_date_text)
    replace_placeholder(ws, "{COLECTARI}", colectari_text)

    # 2) header de tabel (scriem textele vizibile peste placeholders)
    replace_placeholder(ws, "{DENUMIRE SITE}", "Denumire Site")
    replace_placeholder(ws, "{TID}", "TID")
    replace_placeholder(ws, "{DENUMIRE PRODUS}", "Denumire Produs")
    replace_placeholder(ws, "{VALOARE CU TVA}", "Valoare cu TVA")
    replace_placeholder(ws, "{DATA TRANZACTIEI}", "Data Tranzactiei")
    replace_placeholder(ws, "{TOTAL}", "Total")

    # Găsim rândul de start = rândul cu unul din headerele de mai sus (luăm cel pentru „Denumire Site”)
    start_pos = find_cell(ws, "Denumire Site")
    if not start_pos:
        # fallback: încearcă după placeholder-ul original dacă a rămas
        start_pos = find_cell(ws, "{DENUMIRE SITE}")
    if not start_pos:
        # hard fallback: 17, conform șabloanelor tale
        start_row = 17
        start_col = 1
    else:
        start_row, start_col = start_pos

    # Datele încep pe rândul următor
    row = start_row + 1

    # 3) scriem tranzacțiile
    total = 0.0
    for tx in tx_rows:
        site = tx.get("site", "")
        tid = tx.get("tid", "")
        prod = tx.get("produs", "")
        val = float(tx.get("valoare", 0.0))
        dt  = tx.get("data")  # datetime

        ws.cell(row=row, column=start_col + 0, value=site)
        ws.cell(row=row, column=start_col + 1, value=str(tid))
        ws.cell(row=row, column=start_col + 2, value=prod)

        c_val = ws.cell(row=row, column=start_col + 3, value=val)
        c_val.number_format = "0.00"          # <<<<<< 2 zecimale, nu „000”
        total += val

        # data tranzacției: „YYYY-MM-DD HH:MM:SS”
        if isinstance(dt, datetime) and dt != datetime.min:
            ws.cell(row=row, column=start_col + 4, value=dt.strftime("%Y-%m-%d %H:%M:%S"))
        else:
            ws.cell(row=row, column=start_col + 4, value="")

        row += 1

    # 4) total sub ultimele tranzacții
    # găsim celula cu „Total” (am înlocuit placeholderul mai sus)
    total_label_pos = find_cell(ws, "Total")
    if total_label_pos:
        trow, tcol = total_label_pos
        # punem totalul în aceeași coloană ca valorile (start_col + 3)
        total_cell = ws.cell(row=trow, column=start_col + 3, value=total)
        total_cell.number_format = "0.00"     # <<<<<< 2 zecimale

    # Ajustări utile (optional): lățimi consistente ca să nu mai sară textul
    try:
        set_col_width(ws, start_col + 0, 26.0)  # Denumire Site
        set_col_width(ws, start_col + 1, 14.0)  # TID
        set_col_width(ws, start_col + 2, 34.0)  # Denumire Produs
        set_col_width(ws, start_col + 3, 14.0)  # Valoare
        set_col_width(ws, start_col + 4, 22.0)  # Data
    except Exception:
        pass

    # 5) salvează fișierul
    wb.save(out_path)


# ---------------- Orchestrator ----------------

def run_generator(astob_path: str, key_path: str, template_path: str, out_dir: str, out_zip_path: str) -> bool:
    ast = read_table(astob_path)
    key = read_table(key_path)

    # Coloane cheie în ASTOB
    col_site_ast   = find_col(ast, ["Nr. site", "Denumire Site", "Site"])
    col_tid_ast    = find_col(ast, ["Nr. terminal", "TID"])
    col_prod_ast   = find_col(ast, ["Nume Operator", "Denumire Produs"])
    col_sum_ast    = find_col(ast, ["Sumă tranzacție", "Suma tranzactie", "Valoare cu TVA", "Valoare"])
    col_date_ast   = find_col(ast, ["Data tranzacției", "Data tranzactiei", "Data"])
    # oră poate lipsi – nu o fac obligatorie
    try:
        col_time_ast = find_col(ast, ["Ora tranzacției", "Ora tranzactiei", "Ora"])
    except Exception:
        col_time_ast = None

    # Coloane cheie în TABEL CHEIE
    col_client_key = find_col(key, ["DENUMIRE SOCIETATEAGENT", "Denumire societate agent", "Denumire societateagent", "NUME", "CLIENT"])
    col_tid_key    = find_col(key, ["TID", "Nr. terminal"])
    col_cui_key    = find_col(key, ["CUI CNP", "CUI"])
    col_reg_key    = find_col(key, ["NR INREGISTRARE", "Nr. inregistrare R.C."])
    col_addr_key   = find_col(key, ["ADRESA", "Adresa"])

    # normalizăm tipurile
    key["_TID"]    = key[col_tid_key].astype(str).str.strip()
    key["_NUME"]   = key[col_client_key].astype(str).str.strip()
    key["_CUI"]    = key[col_cui_key].astype(str).str.strip()
    key["_REG"]    = key[col_reg_key].astype(str).str.strip()
    key["_ADRESA"] = key[col_addr_key].astype(str).str.strip()

    # map TID -> (nume, cui, reg, adresa, site)
    # „Denumire Site” în Key poate fi absentă; dacă există, o folosim ca SITE.
    site_col_key: Optional[str]
    try:
        site_col_key = find_col(key, ["DENUMIRE SITE", "Denumire Site", "Site"])
        key["_SITE"] = key[site_col_key].astype(str).str.strip()
    except Exception:
        site_col_key = None
        key["_SITE"] = ""

    key_map = {}
    for _, r in key.iterrows():
        tid = str(r["_TID"])
        key_map[tid] = {
            "nume":   r["_NUME"],
            "cui":    r["_CUI"],
            "reg":    r["_REG"],
            "adresa": r["_ADRESA"],
            "site":   r.get("_SITE", ""),
        }

    # Grupăm tranzacțiile din ASTOB pe client (prin TID -> Client din key)
    rows = []
    for _, tr in ast.iterrows():
        tid = str(tr[col_tid_ast]).strip()
        km  = key_map.get(tid)
        if not km:
            # TID necunoscut în tabelul cheie → sare peste
            continue

        client = km["nume"]
        # site din KEY, dacă nu există luăm din ASTOB (coloana „Nr. site” / „Denumire Site”)
        site = km["site"] if km["site"] else str(tr.get(col_site_ast, "")).strip()
        produs = str(tr.get(col_prod_ast, "")).strip()

        val = parse_amount(tr.get(col_sum_ast, 0))
        dt  = parse_datetime(tr.get(col_date_ast, ""), tr.get(col_time_ast) if col_time_ast else None)

        rows.append({
            "client": client,
            "tid": tid,
            "site": site,
            "produs": produs,
            "valoare": val,
            "data": dt,
            "cui": km["cui"],
            "reg": km["reg"],
            "adresa": km["adresa"],
        })

    if not rows:
        # nimic de generat
        with zipfile.ZipFile(out_zip_path, "w", zipfile.ZIP_DEFLATED) as _z:
            pass
        return True

    # Sortăm global după client, apoi dată (cu oră)
    rows.sort(key=lambda r: (r["client"], r["data"]))

    # Data pentru antet = data de azi
    header_date_text = ro_month_upper(datetime.now())

    # Interval „Colectări” = min(data) – max(data) din ASTOB
    dts = [r["data"] for r in rows if isinstance(r["data"], datetime) and r["data"] != datetime.min]
    if dts:
        dmin = min(dts).strftime("%d.%m.%Y")
        dmax = max(dts).strftime("%d.%m.%Y")
        colectari_text = f"{dmin} - {dmax}"
    else:
        colectari_text = ""

    # Creează out_dir
    os.makedirs(out_dir, exist_ok=True)

    # Iterează pe clienți
    ok_any = False
    from itertools import groupby
    for client, grp in groupby(rows, key=lambda r: r["client"]):
        grp_list = list(grp)
        total = sum(r["valoare"] for r in grp_list)
        # Sara peste client dacă TOTAL == 0
        if abs(total) < 1e-9:
            continue

        # info din primul rând pentru antet
        first = grp_list[0]
        client_rc  = first["reg"]
        client_cui = first["cui"]
        client_addr= first["adresa"]

        # Triere tranzacții (data cu oră)
        grp_list.sort(key=lambda r: r["data"])

        # Pregătește listă simplificată pentru scriere în template
        tx_rows = []
        for r in grp_list:
            tx_rows.append({
                "site": r["site"] if r["site"] else client,   # dacă lipsește site, punem numele clientului
                "tid": r["tid"],
                "produs": r["produs"],
                "valoare": r["valoare"],
                "data": r["data"],
            })

        # nume fișier
        safe_client = "".join(ch for ch in str(client) if ch.isalnum() or ch in (" ", "-", "_")).strip()
        out_xlsx = os.path.join(out_dir, f"Ordin - {safe_client}.xlsx")

        fill_workbook_for_client(
            template_path=template_path,
            client_name=client,
            client_rc=client_rc,
            client_cui=client_cui,
            client_addr=client_addr,
            tx_rows=tx_rows,
            colectari_text=colectari_text,
            header_date_text=header_date_text,
            out_path=out_xlsx,
        )
        ok_any = True

    # Ambalează zip
    with zipfile.ZipFile(out_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fn in os.listdir(out_dir):
            full = os.path.join(out_dir, fn)
            if os.path.isfile(full):
                zf.write(full, arcname=fn)

    return ok_any


# ---------------- CLI ----------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--astob", required=True)
    ap.add_argument("--key", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--out-zip", required=True)
    args = ap.parse_args()

    ok = run_generator(args.astob, args.key, args.template, args.out_dir, args.out_zip)
    if not ok:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
